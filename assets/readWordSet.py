import pymysql
from openpyxl import load_workbook
conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', passwd='maxinyin', db='vocabulary', charset='utf8')
cursor = conn.cursor()
wb2 = load_workbook('tofel.xlsx')
sheet = wb2['Sheet2']
for i in range(sheet.max_row):
	row = 'C' + str(i+2);

	sql = "INSERT INTO toefl (word, id) values (\"" + str(sheet[row].value) +"\"," +  str(i+1) + ");";
	print(sql);
	cursor.execute(sql);
	conn.commit()

conn.close()